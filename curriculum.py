import os
import datetime
from openpyxl import Workbook, load_workbook
import requests
from bs4 import BeautifulSoup

def fetch_curriculum(course_code):
    url = f'https://msbte.org.in/DISRESLIVE.aspx?code={course_code}'
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an HTTPError for bad responses

        soup = BeautifulSoup(response.text, 'html.parser')
        curriculum_elements = soup.find_all('div', class_='DISDETTR')
        curriculum_data = [curriculum.text.strip() for curriculum in curriculum_elements]

        return curriculum_data
    except requests.exceptions.RequestException as e:
        print(f"Failed to retrieve data. Error: {e}")
        return None

def download_pdf(pdf_url, save_path):
    try:
        response = requests.get(pdf_url)
        response.raise_for_status()  # Raise an HTTPError for bad responses

        with open(save_path, 'wb') as pdf_file:
            pdf_file.write(response.content)

        print(f"PDF file downloaded successfully and saved as {save_path}")
    except requests.exceptions.RequestException as e:
        print(f"Failed to download PDF. Error: {e}")

def write_to_excel_and_download_pdf(subject_code):
    # Download PDF file
    pdf_url = f'https://msbte.org.in/portal/msbte_files/curriculum_search/papercode_files/{subject_code}.pdf'
    pdf_filename = f'{subject_code}.pdf'
    download_pdf(pdf_url, pdf_filename)

    # Get current time
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Check if Excel file already exists
    excel_filename = "download_records.xlsx"
    if not os.path.exists(excel_filename):
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        # Write headers
        ws.append(["File Name", "Download Time"])
    else:
        # Load existing workbook
        wb = load_workbook(excel_filename)
        ws = wb.active

    # Append file name and download time to the Excel file
    ws.append([pdf_filename, current_time])

    # Save the workbook
    wb.save(excel_filename)

    print(f"Download record written to Excel file: {excel_filename}")

def main():
    # Specify the subject code for the curriculum you want to fetch
    subject_code = input("Enter the subject code for the curriculum: ")

    # Download PDF file and store download record in Excel
    write_to_excel_and_download_pdf(subject_code)

if __name__ == "__main__":
    main()
